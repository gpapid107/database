"""
Export synthetic data from load.sql to Excel.
Each target table becomes a sheet with professional formatting.
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SQL_PATH = "/sessions/eloquent-hopeful-maxwell/mnt/ΒΑΣΕΙΣ ΔΕΔΟΜΕΝΩΝ/hospital-db-project/sql/load.sql"
OUT_PATH = "/sessions/eloquent-hopeful-maxwell/mnt/ΒΑΣΕΙΣ ΔΕΔΟΜΕΝΩΝ/hospital-db-project/data/synthetic_data.xlsx"

# Target tables in desired sheet order
TARGET_TABLES = [
    "STAFF", "Doctor", "Department", "Nurse", "Management",
    "Belongs_Doctor", "Patient", "Patient_Allergy", "Insurance",
    "Operating_Room", "Bed", "Triage", "Admission", "Shift",
    "Shift_Staff", "Medical_Action", "Surgery", "Surgery_Assistant",
    "Exam", "Prescription", "Evaluation", "Doctor_Evaluation", "Entity_Image",
]

# Tables that are NOT synthetic (skip them even if name matches substring)
SKIP_TABLES = {
    "Medicine", "Diagnosis", "KEN", "ICD10", "Substance",
    "Medicine_Composition", "EMA", "MDN",
}


def parse_value_list(s):
    """
    Parse a SQL VALUES row string like: ('foo', 'bar\'s', NULL, 42, 1.5)
    Returns a list of Python values (str, int, float, or None).
    Handles escaped single quotes (\') and backtick-quoted identifiers.
    """
    s = s.strip()
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1]

    values = []
    i = 0
    while i < len(s):
        # skip whitespace
        while i < len(s) and s[i] in " \t":
            i += 1
        if i >= len(s):
            break

        if s[i] == "'":
            # quoted string
            i += 1
            buf = []
            while i < len(s):
                if s[i] == "\\" and i + 1 < len(s) and s[i + 1] == "'":
                    # backslash-escaped quote: \'
                    buf.append("'")
                    i += 2
                elif s[i] == "\\" and i + 1 < len(s) and s[i + 1] == "\\":
                    # backslash-escaped backslash: \\
                    buf.append("\\")
                    i += 2
                elif s[i] == "'" and i + 1 < len(s) and s[i + 1] == "'":
                    # SQL doubled-quote escape: '' -> '
                    buf.append("'")
                    i += 2
                elif s[i] == "'":
                    # End of string
                    i += 1
                    break
                else:
                    buf.append(s[i])
                    i += 1
            values.append("".join(buf))
        else:
            # unquoted: NULL, number, etc.
            j = i
            while j < len(s) and s[j] not in ",":
                j += 1
            token = s[i:j].strip()
            if token.upper() == "NULL":
                values.append(None)
            else:
                try:
                    if "." in token:
                        values.append(float(token))
                    else:
                        values.append(int(token))
                except ValueError:
                    values.append(token)
            i = j

        # skip comma
        while i < len(s) and s[i] in " \t":
            i += 1
        if i < len(s) and s[i] == ",":
            i += 1

    return values


def parse_column_names(col_str):
    """Parse column names from 'Col1, `Col2`, Col3' format."""
    cols = []
    for c in col_str.split(","):
        c = c.strip().strip("`")
        if c:
            cols.append(c)
    return cols


def extract_tables(sql_path, target_tables):
    """
    Read the SQL file and extract INSERT data for target tables.
    Returns dict: table_name -> {"columns": [...], "rows": [[...]]}
    Handles both:
      - Single-row: INSERT INTO `T` (cols) VALUES (row);
      - Multi-row:  INSERT INTO `T` (cols) VALUES\n  (row1),\n  (row2);
    """
    target_set = set(target_tables)
    results = {t: {"columns": None, "rows": []} for t in target_tables}

    insert_re = re.compile(
        r"INSERT INTO `(\w+)`\s*\(([^)]+)\)\s*VALUES\s*(.*)",
        re.IGNORECASE
    )

    with open(sql_path, "r", encoding="utf-8") as f:
        content = f.read()

    # Split on statement boundaries (semicolon at end of line or followed by newline)
    # We process statement by statement
    # Strategy: find all INSERT INTO blocks for target tables
    # Use a streaming approach to handle multi-line VALUES

    # Split into logical statements by finding INSERT INTO `TARGET` patterns
    # and collecting until the semicolon

    lines = content.split("\n")
    i = 0
    total = len(lines)

    while i < total:
        line = lines[i]

        m = re.match(r"\s*INSERT INTO `(\w+)`\s*\(([^)]+)\)\s*VALUES\s*(.*)", line, re.IGNORECASE)
        if not m:
            i += 1
            continue

        table_name = m.group(1)
        col_str = m.group(2)
        rest = m.group(3).strip()

        if table_name not in target_set:
            i += 1
            continue

        columns = parse_column_names(col_str)

        # Collect the full VALUES block until we hit the terminating semicolon
        # rest may already contain part of it
        # Accumulate lines until we see a semicolon at end (after closing paren)
        block = rest
        if not block.endswith(";"):
            i += 1
            while i < total:
                next_line = lines[i].strip()
                block += " " + next_line
                i += 1
                if next_line.endswith(";"):
                    break
        else:
            i += 1

        # Remove trailing semicolon
        block = block.rstrip().rstrip(";").strip()

        # Now parse individual value rows: find each (...) group
        # We need to split on ")," or ");" but respect quoted strings
        rows = []
        pos = 0
        blen = len(block)

        while pos < blen:
            # skip whitespace and commas
            while pos < blen and block[pos] in " \t\n\r,":
                pos += 1
            if pos >= blen:
                break
            if block[pos] != "(":
                pos += 1
                continue
            # find matching closing paren, respecting quoted strings
            start = pos
            pos += 1
            depth = 1
            in_quote = False
            while pos < blen and depth > 0:
                c = block[pos]
                if in_quote:
                    if c == "\\" and pos + 1 < blen and block[pos + 1] in ("'", "\\"):
                        # backslash escape: \' or \\
                        pos += 2
                        continue
                    if c == "'" and pos + 1 < blen and block[pos + 1] == "'":
                        # SQL doubled-quote escape '' inside string — skip both
                        pos += 2
                        continue
                    if c == "'":
                        in_quote = False
                else:
                    if c == "'":
                        in_quote = True
                    elif c == "(":
                        depth += 1
                    elif c == ")":
                        depth -= 1
                pos += 1
            row_str = block[start:pos]
            values = parse_value_list(row_str)
            if values:
                rows.append(values)

        if results[table_name]["columns"] is None:
            results[table_name]["columns"] = columns
        results[table_name]["rows"].extend(rows)

    return results


def style_sheet(ws, columns, rows):
    HEADER_FONT = Font(name="Arial", size=11, bold=True)
    DATA_FONT = Font(name="Arial", size=10)
    HEADER_FILL = PatternFill("solid", fgColor="C8DCF0")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Write headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT

    # Auto-filter
    if columns:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}1"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-fit column widths
    col_widths = [len(str(c)) for c in columns]
    for row in rows:
        for ci, val in enumerate(row):
            if ci < len(col_widths) and val is not None:
                col_widths[ci] = max(col_widths[ci], min(len(str(val)), 60))

    for ci, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width + 2

    # Row height for header
    ws.row_dimensions[1].height = 20


def main():
    print("Parsing SQL file...")
    data = extract_tables(SQL_PATH, TARGET_TABLES)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for table_name in TARGET_TABLES:
        info = data[table_name]
        columns = info["columns"]
        rows = info["rows"]

        if columns is None:
            print(f"  WARNING: No data found for {table_name}")
            columns = []

        sheet_name = table_name[:31]  # Excel limit
        ws = wb.create_sheet(title=sheet_name)
        style_sheet(ws, columns, rows)
        print(f"  {table_name}: {len(rows)} rows, {len(columns)} columns")

    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")


if __name__ == "__main__":
    main()
