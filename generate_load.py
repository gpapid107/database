#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_load.py

Generator για sql/load.sql για το schema του Γενικού Νοσοκομείου «Υγειόπολης».

Χρήση:
  python3 generate_load.py --output load.sql --seed 23878
  python3 generate_load.py --output load.sql --ema-xlsx article-57-product-data_en.xlsx

Σημειώσεις:
- Από προεπιλογή δημιουργεί ΠΛΗΡΕΣ ημερολόγιο εφημεριών: κάθε ημέρα, κάθε τμήμα, και οι 3 βάρδιες, με πλήρη στελέχωση Shift_Staff.
- Δημιουργεί 1000 admissions με ημερομηνίες από 2024-01-01 έως σήμερα.
- Δημιουργεί συνεπή prescription/evaluation/doctor_evaluation δεδομένα: Doctor_Evaluation μπαίνει μόνο
  για γιατρούς που όντως συνταγογράφησαν στη συγκεκριμένη νοσηλεία.
- Για τις βάρδιες, το script αυξάνει το συνθετικό προσωπικό αρκετά ώστε να καλύπτεται πλήρως το 24/7 πρόγραμμα χωρίς παραβίαση των μηνιαίων ορίων ή του 8ώρου ανάπαυσης.
- Το install/triggers schema έχει circular load issue στο Department director trigger:
  Department BEFORE INSERT ζητά να υπάρχει ήδη Belongs_Doctor για το ίδιο Department, ενώ το
  Belongs_Doctor έχει FK προς Department. Για αυτό το load.sql βάζει προσωρινά FOREIGN_KEY_CHECKS=0
  μέχρι να φορτωθούν Belongs_Doctor και Department. Η τελική κατάσταση είναι συνεπής.
"""

from __future__ import annotations

import argparse
import csv
import random
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable


DEPARTMENTS = [
    'Καρδιολογία', 'Χειρουργική', 'ΜΕΘ', 'Επείγοντα', 'Νευρολογία',
    'Παθολογία', 'Ορθοπεδική', 'Ουρολογία', 'Οφθαλμολογία', 'ΩΡΛ',
    'Πνευμονολογία', 'Γαστρεντερολογία', 'Νεφρολογία', 'Αιματολογία', 'Ογκολογία'
]

SPECIALTIES = [
    'Καρδιολογία', 'Χειρουργική', 'Εντατικολογία', 'Επείγουσα Ιατρική', 'Νευρολογία',
    'Παθολογία', 'Ορθοπεδική', 'Ουρολογία', 'Οφθαλμολογία', 'Ωτορινολαρυγγολογία',
    'Πνευμονολογία', 'Γαστρεντερολογία', 'Νεφρολογία', 'Αιματολογία', 'Ογκολογία'
]

FIRST_NAMES_M = ['Γιώργος', 'Νίκος', 'Κώστας', 'Δημήτρης', 'Αλέξανδρος', 'Παναγιώτης', 'Ανδρέας', 'Μιχάλης', 'Στέφανος', 'Βασίλης']
FIRST_NAMES_F = ['Μαρία', 'Ελένη', 'Άννα', 'Κατερίνα', 'Σοφία', 'Ιωάννα', 'Δήμητρα', 'Αναστασία', 'Χριστίνα', 'Γεωργία']
LAST_NAMES = ['Παπαδόπουλος', 'Γεωργίου', 'Νικολάου', 'Δημητρίου', 'Ιωάννου', 'Αθανασίου', 'Κωνσταντίνου', 'Πέτρου', 'Αντωνίου', 'Σταύρου', 'Οικονόμου', 'Μακρής']
FATHER_NAMES = ['Ιωάννης', 'Νικόλαος', 'Δημήτριος', 'Γεώργιος', 'Κωνσταντίνος', 'Παναγιώτης']
INSURANCES = ['ΕΦΚΑ', 'ΕΟΠΥΥ', 'Ιδιωτική Ασφάλεια', 'Ανασφάλιστος']
NURSE_RANKS = ['Βοηθός Νοσηλευτή', 'Νοσηλευτής', 'Προϊστάμενος']
MANAGEMENT_ROLES = ['Γραμματέας', 'Λογιστής', 'Ανθρώπινο Δυναμικό']
BED_TYPES = ['ΜΕΘ', 'Μονόκλινο', 'Πολύκλινο']
SYMPTOMS = ['Πυρετός', 'Θωρακικό άλγος', 'Δύσπνοια', 'Κοιλιακό άλγος', 'Κάταγμα', 'Ζάλη', 'Αιμορραγία', 'Κεφαλαλγία']
EXAM_TYPES = ['Αιματολογική', 'Βιοχημική', 'Ακτινογραφία', 'Υπέρηχος', 'CT', 'MRI']

# Πραγματικοί/τυπικοί ICD-10 κωδικοί και περιγραφές, όχι εφευρεμένοι κωδικοί.
DIAGNOSES = [
    ('I21.9', 'Οξύ έμφραγμα του μυοκαρδίου, μη καθορισμένο'),
    ('I50.9', 'Καρδιακή ανεπάρκεια, μη καθορισμένη'),
    ('J18.9', 'Πνευμονία, μη καθορισμένη'),
    ('J44.9', 'Χρόνια αποφρακτική πνευμονοπάθεια, μη καθορισμένη'),
    ('K35.8', 'Οξεία σκωληκοειδίτιδα, άλλη/μη καθορισμένη'),
    ('K80.2', 'Χολολιθίαση χωρίς χολοκυστίτιδα'),
    ('N18.9', 'Χρόνια νεφρική νόσος, μη καθορισμένη'),
    ('E11.9', 'Σακχαρώδης διαβήτης τύπου 2 χωρίς επιπλοκές'),
    ('S72.0', 'Κάταγμα αυχένα μηριαίου'),
    ('G45.9', 'Παροδικό εγκεφαλικό ισχαιμικό επεισόδιο, μη καθορισμένο'),
    ('C34.9', 'Κακοήθες νεόπλασμα βρόγχου ή πνεύμονα, μη καθορισμένο'),
    ('D64.9', 'Αναιμία, μη καθορισμένη'),
    ('H26.9', 'Καταρράκτης, μη καθορισμένος'),
    ('N20.0', 'Λίθος νεφρού'),
    ('C50.9', 'Κακοήθες νεόπλασμα μαστού, μη καθορισμένο'),
]

# Κ.Ε.Ν. κωδικοί από το παρεχόμενο αρχείο ΚΕΝ/εκφώνηση. Κρατάμε μικρό αλλά πραγματικό reference subset.
KEN_ROWS = [
    ('Ν05Α', 439.00, 1),
    ('Ν30Χ', 1191.00, 6),
    ('Α22Χ', 630.00, 3),
    ('Α25Χ', 949.00, 5),
    ('Κ06Χ', 5077.00, 7),
    ('Κ15Χ', 2013.00, 2),
    ('Η07Χ', 2037.00, 6),
    ('Η08Χ', 1273.00, 2),
    ('Ο16Α', 592.00, 1),
    ('Ω31Α', 444.00, 2),
    ('Φ61Α', 280.00, 1),
    ('Ν39Χ', 537.00, 7),
    ('Α36Α', 2412.00, 14),
    ('Κ20Α', 760.00, 1),
    ('W10Α', 2014.00, 3),
]

# Fallback medicines. Αν δοθεί --ema-xlsx, χρησιμοποιούνται δεδομένα από EMA Article 57.
FALLBACK_MEDICINES = [
    ('EMA-FALLBACK-001', 'Paracetamol 500mg tablet', ['Paracetamol']),
    ('EMA-FALLBACK-002', 'Amoxicillin 500mg capsule', ['Amoxicillin']),
    ('EMA-FALLBACK-003', 'Ibuprofen 400mg tablet', ['Ibuprofen']),
    ('EMA-FALLBACK-004', 'Metformin 850mg tablet', ['Metformin']),
    ('EMA-FALLBACK-005', 'Atorvastatin 20mg tablet', ['Atorvastatin']),
    ('EMA-FALLBACK-006', 'Omeprazole 20mg capsule', ['Omeprazole']),
    ('EMA-FALLBACK-007', 'Amlodipine 5mg tablet', ['Amlodipine']),
    ('EMA-FALLBACK-008', 'Salbutamol inhaler', ['Salbutamol']),
    ('EMA-FALLBACK-009', 'Cefuroxime 500mg tablet', ['Cefuroxime']),
    ('EMA-FALLBACK-010', 'Insulin glargine injection', ['Insulin glargine']),
    ('EMA-FALLBACK-011', 'Clopidogrel 75mg tablet', ['Clopidogrel']),
    ('EMA-FALLBACK-012', 'Furosemide 40mg tablet', ['Furosemide']),
]


@dataclass(frozen=True)
class StaffMember:
    amka: str
    kind: str
    first: str
    last: str
    department: str | None = None
    rank: str | None = None
    specialty: str | None = None


def q(value: Any) -> str:
    if value is None:
        return 'NULL'
    if isinstance(value, bool):
        return '1' if value else '0'
    if isinstance(value, (int, float)):
        return str(value)
    # Το datetime πρέπει να ελεγχθεί πριν από το date.
    if isinstance(value, datetime):
        return "'" + value.isoformat(sep=' ') + "'"
    if isinstance(value, date):
        return "'" + value.isoformat() + "'"
    s = str(value).replace(chr(92), chr(92) + chr(92)).replace("'", "''")
    return "'" + s + "'"


def row(values: Iterable[Any]) -> str:
    return '(' + ', '.join(q(v) for v in values) + ')'


class SQLWriter:
    def __init__(self) -> None:
        self.lines: list[str] = []

    def raw(self, text: str = '') -> None:
        self.lines.append(text)

    def insert_many(self, table: str, columns: list[str], rows: list[tuple[Any, ...]], batch: int = 500) -> None:
        if not rows:
            return
        cols = ', '.join(f'`{c}`' for c in columns)
        for i in range(0, len(rows), batch):
            chunk = rows[i:i + batch]
            self.lines.append(f"INSERT INTO `{table}` ({cols}) VALUES")
            self.lines.append(',\n'.join(row(r) for r in chunk) + ';')

    def text(self) -> str:
        return '\n'.join(self.lines) + '\n'


def amka(prefix: int, i: int) -> str:
    # 11 ψηφία ως string.
    return f"{prefix}{i:08d}"[:11]


def phone(i: int) -> str:
    return f"69{10000000 + i:08d}"[:10]


def email(prefix: str, i: int) -> str:
    return f"{prefix}{i}@ygeiopolis.gr"


def rand_name(rng: random.Random, female: bool | None = None) -> tuple[str, str]:
    if female is None:
        female = rng.random() < 0.5
    first = rng.choice(FIRST_NAMES_F if female else FIRST_NAMES_M)
    last = rng.choice(LAST_NAMES)
    return first, last


def daterange(d0: date, d1: date) -> Iterable[date]:
    cur = d0
    while cur <= d1:
        yield cur
        cur += timedelta(days=1)


def random_date(rng: random.Random, start: date, end: date) -> date:
    days = (end - start).days
    return start + timedelta(days=rng.randint(0, max(days, 0)))


def load_ema_xlsx(path: Path, limit: int | None = None) -> list[tuple[str, str, list[str]]]:
    """Best-effort parser για EMA Article 57 xlsx.

    Φορτώνει ΟΛΕΣ τις γραμμές του αρχείου, εκτός αν δοθεί limit.
    Ψάχνει ευέλικτα στήλες με Product number/code, Product name και Active substance(s).
    Αν αποτύχει, επιστρέφει fallback demo rows.
    """
    if not path or not path.exists():
        return FALLBACK_MEDICINES
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return FALLBACK_MEDICINES

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return FALLBACK_MEDICINES

    header_idx = 0
    best_score = -1
    for idx, r in enumerate(rows[:20]):
        vals = [str(x).strip().lower() if x is not None else '' for x in r]
        score = sum(any(k in v for v in vals) for k in ['product', 'active', 'substance', 'name'])
        if score > best_score:
            best_score = score
            header_idx = idx

    headers = [str(x).strip().lower() if x is not None else '' for x in rows[header_idx]]

    def find_col(*needles: str) -> int | None:
        for j, h in enumerate(headers):
            if all(n in h for n in needles):
                return j
        return None

    code_col = find_col('product', 'number') or find_col('product', 'code') or find_col('ev', 'code') or 0
    name_col = find_col('product', 'name') or find_col('name') or 1
    subs_col = find_col('active', 'substance') or find_col('substance')
    if subs_col is None:
        return FALLBACK_MEDICINES

    out: list[tuple[str, str, list[str]]] = []
    seen = set()
    for r in rows[header_idx + 1:]:
        code = str(r[code_col]).strip() if code_col < len(r) and r[code_col] else ''
        name = str(r[name_col]).strip() if name_col < len(r) and r[name_col] else ''
        subs_raw = str(r[subs_col]).strip() if subs_col < len(r) and r[subs_col] else ''
        if not code or not name or not subs_raw:
            continue
        # Το αρχείο EMA συχνά διαχωρίζει με | τις δραστικές ουσίες.
        substances = [s.strip() for s in re.split(r'\s*\|\s*|\s*;\s*|\s*,\s*', subs_raw) if s.strip()]
        if not substances:
            continue
        # Το EMA Article 57 δεν έχει πάντα κατάλληλο μοναδικό πεδίο για το δικό μας schema.
        # Σε κάποια exports το auto-detected code βγαίνει π.χ. 'A' ή επαναλαμβανόμενο όνομα.
        # Άρα κρατάμε ΟΛΕΣ τις γραμμές, αλλά φτιάχνουμε δικό μας μοναδικό EMA_Code.
        final_code = f'EMA{len(out) + 1:08d}'
        seen.add(final_code)
        out.append((final_code, name[:200], substances[:4]))
        if limit is not None and len(out) >= limit:
            break
    return out or FALLBACK_MEDICINES


def autodiscover_file(reference_dir: Path, keywords: list[str], suffixes: tuple[str, ...]) -> Path | None:
    candidates = []
    if not reference_dir.exists():
        return None
    for p in reference_dir.iterdir():
        if not p.is_file() or p.suffix.lower() not in suffixes:
            continue
        lname = p.name.lower()
        score = sum(1 for k in keywords if k.lower() in lname)
        if score:
            candidates.append((score, p))
    candidates.sort(key=lambda x: (-x[0], x[1].name))
    return candidates[0][1] if candidates else None


def read_legacy_doc_text(path: Path) -> str:
    if not path or not path.exists():
        return ''
    suffix = path.suffix.lower()
    if suffix in {'.txt', '.csv'}:
        return path.read_text(encoding='utf-8', errors='ignore')
    if suffix == '.docx':
        try:
            import docx  # type: ignore
            doc = docx.Document(str(path))
            return chr(10).join(p.text for p in doc.paragraphs)
        except Exception:
            pass
    import subprocess
    import tempfile
    if suffix == '.doc':
        # Πολλά παλιά .doc του Υπ. Υγείας περιέχουν το κείμενο σε UTF-16LE μέσα στο OLE file.
        # antiword/LibreOffice συχνά το βγάζουν ως ?????, οπότε πρώτα κάνουμε raw UTF-16LE extraction.
        try:
            raw_text = path.read_bytes().decode('utf-16le', errors='ignore')
            greek_hits = sum(raw_text.count(x) for x in ['ΚΕΝ', 'ΚΩΔΙΚ', 'ΜΔΝ', 'Νοσήλια', 'Μεταμόσχευση'])
            if greek_hits >= 2:
                return raw_text.replace('\x07', ' ').replace('\x00', ' ')
        except Exception:
            pass
        try:
            with tempfile.TemporaryDirectory() as td:
                out = Path(td) / (path.stem + '.txt')
                subprocess.run(['textutil', '-convert', 'txt', str(path), '-output', str(out)], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                if out.exists():
                    return out.read_text(encoding='utf-8', errors='ignore')
        except Exception:
            pass
        for cmd in (['antiword', str(path)], ['catdoc', str(path)]):
            try:
                res = subprocess.run(cmd, check=True, capture_output=True)
                return res.stdout.decode('utf-8', errors='ignore')
            except Exception:
                pass
        try:
            with tempfile.TemporaryDirectory() as td:
                subprocess.run(['soffice', '--headless', '--convert-to', 'txt:Text', '--outdir', td, str(path)], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                txts = list(Path(td).glob('*.txt'))
                if txts:
                    return txts[0].read_text(encoding='utf-8', errors='ignore')
        except Exception:
            pass
    return path.read_bytes().decode('utf-8', errors='ignore')


def parse_euro_amount(s: str) -> float | None:
    cleaned = s.replace(chr(160), ' ').replace('€', '').strip()
    token = None
    for part in cleaned.split():
        p = part.strip()
        if any(ch.isdigit() for ch in p):
            token = ''.join(ch for ch in p if ch.isdigit() or ch in '.,')
            break
    if not token:
        return None
    try:
        return float(token.replace('.', '').replace(',', '.'))
    except ValueError:
        return None


def load_ken_reference(path: Path | None) -> list[tuple[str, float, int]]:
    if not path:
        return KEN_ROWS
    text = read_legacy_doc_text(path)
    norm = ' '.join(text.replace(chr(160), ' ').replace(chr(9), ' ').split())

    greek_letters = 'ΑΒΓΔΕΖΗΘΙΚΛΜΝΞΟΠΡΣΤΥΦΧΨΩ'
    latin_letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    letters = greek_letters + latin_letters
    code_pat = '[' + letters + '][0-9]{2}[' + letters + '](?:[a-zα-ω])?'
    amount_pat = '([0-9]{1,3}(?:[.][0-9]{3})*(?:,[0-9]+)?|[0-9]+(?:,[0-9]+)?)'

    out: list[tuple[str, float, int]] = []
    seen = set()

    pattern = re.compile('(' + code_pat + ')' + '.{0,260}?' + amount_pat + '[ ]*€[ ]*([0-9]{1,3})')
    for m in pattern.finditer(norm):
        code = m.group(1).strip()
        if 'ΤΚΑ' in code or code in seen:
            continue
        cost = parse_euro_amount(m.group(2))
        try:
            mdn = int(m.group(3))
        except ValueError:
            continue
        if cost is not None and cost > 0 and mdn > 0:
            seen.add(code)
            out.append((code[:10], round(cost, 2), mdn))

    if not out:
        loose = re.compile('(' + code_pat + ')' + '.{0,260}?' + '([0-9]{1,3}(?:[.][0-9]{3})+|[0-9]{3,6})' + '[ ]+([0-9]{1,3})')
        for m in loose.finditer(norm):
            code = m.group(1).strip()
            if 'ΤΚΑ' in code or code in seen:
                continue
            cost = parse_euro_amount(m.group(2))
            try:
                mdn = int(m.group(3))
            except ValueError:
                continue
            if cost is not None and cost > 0 and mdn > 0:
                seen.add(code)
                out.append((code[:10], round(cost, 2), mdn))

    if len(out) <= len(KEN_ROWS):
        print('WARNING: KEN parser found only', len(out), 'rows from', path)
        print('Convert the .doc to .txt and pass --ken-doc that .txt file if needed.')
    return out or KEN_ROWS


def spreadsheet_rows(path: Path) -> list[list[str]]:
    if not path or not path.exists():
        return []
    if path.suffix.lower() == '.csv':
        with path.open(newline='', encoding='utf-8', errors='ignore') as f:
            return [[str(c or '').strip() for c in r] for r in csv.reader(f)]

    def read_with_pandas(p: Path) -> list[list[str]]:
        import pandas as pd  # type: ignore
        sheets = pd.read_excel(p, sheet_name=None, dtype=str, header=None)
        out_rows: list[list[str]] = []
        for df in sheets.values():
            for vals in df.fillna('').values.tolist():
                out_rows.append([str(v).strip() for v in vals])
        return out_rows

    try:
        return read_with_pandas(path)
    except Exception as first_exc:
        # Αν είναι παλιό .xls και λείπει xlrd, δοκιμάζουμε αυτόματη μετατροπή με LibreOffice.
        if path.suffix.lower() == '.xls':
            import subprocess
            import tempfile
            try:
                with tempfile.TemporaryDirectory() as td:
                    subprocess.run(
                        ['soffice', '--headless', '--convert-to', 'xlsx', '--outdir', td, str(path)],
                        check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                    )
                    converted = Path(td) / (path.stem + '.xlsx')
                    if converted.exists():
                        return read_with_pandas(converted)
            except Exception:
                pass
        raise SystemExit(f'Δεν μπόρεσα να διαβάσω το Excel {path}. Για .xls εγκατέστησε xlrd ή LibreOffice.') from first_exc


def load_icd_reference(paths: list[Path]) -> list[tuple[str, str]]:
    found: dict[str, str] = {}

    def is_icd(code: str) -> bool:
        if len(code) < 3 or len(code) > 10:
            return False
        if not (code[0].isalpha() and code[1:3].isdigit()):
            return False
        rest = code[3:]
        if rest.startswith('.'):
            rest = rest[1:]
        return all(ch.isalnum() for ch in rest)

    for path in paths:
        if not path.exists():
            continue
        if path.suffix.lower() in {'.xls', '.xlsx', '.csv'}:
            rows = spreadsheet_rows(path)
        else:
            rows = [[x] for x in read_legacy_doc_text(path).splitlines()]
        for r in rows:
            cells = [str(c).strip() for c in r if str(c).strip()]
            for idx, cell in enumerate(cells):
                code = cell.upper().strip()
                if is_icd(code):
                    desc = cells[idx + 1] if idx + 1 < len(cells) else (cells[-1] if len(cells) > 1 else '')
                    if desc and not is_icd(desc.upper()) and len(desc) > 2:
                        found.setdefault(code[:10], desc[:200])
    return sorted(found.items()) or DIAGNOSES


def load_medical_action_catalog(paths: list[Path]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen = set()
    for path in paths:
        if not path.exists() or path.suffix.lower() not in {'.xls', '.xlsx', '.csv'}:
            continue
        for r in spreadsheet_rows(path):
            cells = [str(c).strip() for c in r if str(c).strip()]
            if len(cells) < 2:
                continue
            joined = ' | '.join(cells)
            lower = joined.lower()
            if any(h in lower for h in ['κωδ', 'περιγραφ', 'ονομασία', 'τιμή', 'κόστος']):
                continue
            candidates = [c for c in cells if len(c) >= 6 and any('α' <= ch.lower() <= 'ω' for ch in c)]
            if not candidates:
                continue
            name = max(candidates, key=len)[:400]
            key = name.lower()
            if key in seen:
                continue
            seen.add(key)
            low_name = name.lower()
            if any(w in low_name for w in ['χειρουργ', 'επέμβ', 'επεμβ', 'τομή', 'αφαίρεση', 'λαπαροσκοπ']):
                typ = 'Χειρουργική'
            elif any(w in low_name for w in ['διάγνω', 'βιοψ', 'απεικον', 'ακτινο', 'ενδοσκόπ']):
                typ = 'Διαγνωστική'
            else:
                typ = 'Θεραπευτική'
            out.append({'name': name, 'type': typ})
    return out


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument('--output', default='load.sql')
    parser.add_argument('--seed', type=int, default=23878)
    parser.add_argument('--start-date', default='2024-01-01')
    parser.add_argument('--end-date', default=date.today().isoformat())
    parser.add_argument('--reference-dir', default='.', help='Φάκελος με τα reference αρχεία της εκφώνησης')
    parser.add_argument('--ema-xlsx', default=None, help='EMA Article 57 xlsx. Αν δεν δοθεί, γίνεται auto-discovery στον reference-dir.')
    parser.add_argument('--ken-doc', default=None, help='Αρχείο ΚΕΝ .doc/.txt. Αν δεν δοθεί, γίνεται auto-discovery στον reference-dir.')
    parser.add_argument('--icd-file', action='append', default=[], help='ICD-10 αρχείο .xls/.xlsx/.csv/.doc/.txt. Μπορεί να δοθεί πολλές φορές. Αν δεν δοθεί, σαρώνει τα reference αρχεία.')
    parser.add_argument('--actions-xls', action='append', default=[], help='Αρχείο ιατρικών πράξεων .xls/.xlsx/.csv. Μπορεί να δοθεί πολλές φορές. Αν δεν δοθεί, γίνεται auto-discovery.')
    parser.add_argument('--full-shift-calendar', action='store_true', help='Συμβατότητα με παλιότερη έκδοση: πλέον το πλήρες, πλήρως στελεχωμένο ημερολόγιο βαρδιών παράγεται από προεπιλογή.')
    parser.add_argument('--strict-reference', action='store_true', help='Αν λείπει ή δεν διαβαστεί κάποιο reference αρχείο, σταματάει αντί να χρησιμοποιήσει fallback subset.')
    args = parser.parse_args()

    reference_dir = Path(args.reference_dir)
    ema_path = Path(args.ema_xlsx) if args.ema_xlsx else autodiscover_file(reference_dir, ['article-57', 'ema', 'product-data'], ('.xlsx', '.xls'))
    ken_path = Path(args.ken_doc) if args.ken_doc else autodiscover_file(reference_dir, ['κλειστ', 'ken', 'νοσηλ', '4.1'], ('.doc', '.docx', '.txt'))
    action_paths = [Path(p) for p in args.actions_xls]
    if not action_paths:
        action_paths = [
            p for p in reference_dir.iterdir()
            if p.is_file()
            and p.suffix.lower() in {'.xls', '.xlsx', '.csv'}
            and not any(k in p.name.lower() for k in ['article-57', 'product-data', 'ema', 'icd'])
        ]
    icd_paths = [Path(p) for p in args.icd_file]
    if not icd_paths:
        # Μην ταΐζουμε στον ICD parser τα ΚΕΝ/ιατρικές πράξεις, γιατί μπορεί να βρει
        # τυχαία strings που μοιάζουν με ICD codes. Για auto-discovery κρατάμε μόνο αρχεία
        # που στο όνομα δείχνουν ICD. Αν δεν υπάρχει τέτοιο, κάνει fallback στο παλιό scan.
        icd_paths = [
            p for p in reference_dir.iterdir()
            if p.is_file()
            and p.suffix.lower() in {'.xls', '.xlsx', '.csv', '.doc', '.docx', '.txt'}
            and 'icd' in p.name.lower()
        ]
        if not icd_paths:
            icd_paths = [p for p in reference_dir.iterdir() if p.is_file() and p.suffix.lower() in {'.xls', '.xlsx', '.csv', '.doc', '.docx', '.txt'}]

    rng = random.Random(args.seed)
    start_date = date.fromisoformat(args.start_date)
    end_date = date.fromisoformat(args.end_date)
    if end_date < start_date:
        raise ValueError('end-date πριν από start-date')

    writer = SQLWriter()
    writer.raw('-- Auto-generated load.sql')
    writer.raw(f'-- Generated by generate_load.py with seed={args.seed}')
    writer.raw('USE `mydb`;')
    writer.raw('SET NAMES utf8mb4;')
    writer.raw('SET @OLD_FOREIGN_KEY_CHECKS=@@FOREIGN_KEY_CHECKS;')
    writer.raw('SET FOREIGN_KEY_CHECKS=0;')
    writer.raw('SET @OLD_UNIQUE_CHECKS=@@UNIQUE_CHECKS;')
    writer.raw('SET UNIQUE_CHECKS=0;')
    writer.raw('')

    # -------------------- Staff / Doctors --------------------
    staff_rows: list[tuple[Any, ...]] = []
    doctor_rows: list[tuple[Any, ...]] = []
    doctor_members: list[StaffMember] = []
    director_by_dept: dict[str, str] = {}

    # 15 διευθυντές, ένας ανά τμήμα.
    for i, dept in enumerate(DEPARTMENTS, 1):
        first, last = rand_name(rng)
        a = amka(100, i)
        staff_rows.append((a, first, last, rng.randint(46, 64), email('doctor', i), phone(i), random_date(rng, date(2000, 1, 1), date(2023, 12, 31)), 'Ιατρός'))
        doctor_rows.append((f'LIC-DIR-{i:04d}', SPECIALTIES[i - 1], 'Διευθυντής', a, None))
        doctor_members.append(StaffMember(a, 'Ιατρός', first, last, dept, 'Διευθυντής', SPECIALTIES[i - 1]))
        director_by_dept[dept] = a

    # Επιμελητές Α/Β και ειδικευόμενοι. Οι ειδικευόμενοι παίρνουν senior supervisor.
    # Κρατάμε πολύ πάνω από το ελάχιστο των 80 ιατρών, γιατί πλήρης καθημερινή κάλυψη
    # 15 τμημάτων × 3 βαρδιών × 3 ιατρών απαιτεί περίπου 4.050 ιατρικές βάρδιες/μήνα.
    # Με όριο 15 βάρδιες/μήνα ανά ιατρό χρειάζονται τουλάχιστον ~270 ιατροί.
    ranks_plan = ['Επιμελητής Α'] * 140 + ['Επιμελητής Β'] * 220 + ['Ειδικευόμενος'] * 55
    senior_pool = [d.amka for d in doctor_members]
    for j, rank in enumerate(ranks_plan, 16):
        dept = DEPARTMENTS[(j - 1) % len(DEPARTMENTS)]
        spec = SPECIALTIES[(j - 1) % len(SPECIALTIES)]
        first, last = rand_name(rng)
        a = amka(100, j)
        age = rng.randint(27, 34) if rank == 'Ειδικευόμενος' else rng.randint(35, 58)
        staff_rows.append((a, first, last, age, email('doctor', j), phone(j), random_date(rng, date(2010, 1, 1), end_date), 'Ιατρός'))
        supervisor = rng.choice(senior_pool) if rank == 'Ειδικευόμενος' else (director_by_dept[dept] if rng.random() < 0.35 else None)
        doctor_rows.append((f'LIC-{j:05d}', spec, rank, a, supervisor))
        doctor_members.append(StaffMember(a, 'Ιατρός', first, last, dept, rank, spec))
        if rank in ('Επιμελητής Α', 'Διευθυντής'):
            senior_pool.append(a)

    # Nurses και management.
    nurse_members: list[StaffMember] = []
    management_members: list[StaffMember] = []
    nurse_rows: list[tuple[Any, ...]] = []
    management_rows: list[tuple[Any, ...]] = []

    for i in range(1, 451):
        dept = DEPARTMENTS[(i - 1) % len(DEPARTMENTS)]
        first, last = rand_name(rng, female=rng.random() < 0.65)
        a = amka(200, i)
        staff_rows.append((a, first, last, rng.randint(24, 61), email('nurse', i), phone(1000 + i), random_date(rng, date(2005, 1, 1), end_date), 'Νοσηλευτής'))
        rank = rng.choices(NURSE_RANKS, weights=[2, 7, 1], k=1)[0]
        nurse_rows.append((rank, a, dept))
        nurse_members.append(StaffMember(a, 'Νοσηλευτής', first, last, dept, rank))

    for i in range(1, 141):
        dept = DEPARTMENTS[(i - 1) % len(DEPARTMENTS)]
        first, last = rand_name(rng)
        a = amka(300, i)
        staff_rows.append((a, first, last, rng.randint(25, 64), email('admin', i), phone(2000 + i), random_date(rng, date(2000, 1, 1), end_date), 'Διοικητικός'))
        role = rng.choice(MANAGEMENT_ROLES)
        office = f'{dept[:3]}-{100 + i}'
        management_rows.append((role, office, a, dept))
        management_members.append(StaffMember(a, 'Διοικητικός', first, last, dept, role))

    writer.insert_many('STAFF', ['Staff_AMKA', 'First_Name', 'Last_Name', 'Age', 'Email', 'Phone_Number', 'Hiring_Date', 'Staff_Type'], staff_rows)
    writer.insert_many('Doctor', ['License_Number', 'Specialty', 'Rank', 'Staff_AMKA', 'Supervisor_AMKA'], doctor_rows)

    # Belongs_Doctor πριν από Department λόγω του trigger department_director_*.
    belongs_rows: list[tuple[Any, ...]] = []
    for d in doctor_members:
        belongs_rows.append((d.amka, d.department))
        # Μικρή πιθανότητα δεύτερου τμήματος για πιο πλούσια queries.
        if rng.random() < 0.18:
            extra = rng.choice([x for x in DEPARTMENTS if x != d.department])
            belongs_rows.append((d.amka, extra))
    # Εξασφαλίζουμε ότι οι διευθυντές ανήκουν στο τμήμα που διευθύνουν.
    for dept, dir_amka in director_by_dept.items():
        if (dir_amka, dept) not in belongs_rows:
            belongs_rows.append((dir_amka, dept))
    writer.insert_many('Belongs_Doctor', ['Doctor_AMKA', 'Department_Name'], sorted(set(belongs_rows)))

    department_rows = []
    for i, dept in enumerate(DEPARTMENTS, 1):
        department_rows.append((dept, f'Τμήμα {dept}', 20, (i - 1) % 6, 'Κεντρικό' if i <= 8 else 'Πτέρυγα Β', director_by_dept[dept]))
    writer.insert_many('Department', ['Department_Name', 'Description', 'Num_of_Beds', 'Floor', 'Building', 'DirectedBy'], department_rows)

    writer.insert_many('Nurse', ['Nurse_Rank', 'Staff_AMKA', 'NurseBelongsDepartment'], nurse_rows)
    writer.insert_many('Management', ['Role', 'Office', 'Staff_AMKA', 'ManagerBelongsDepartment'], management_rows)
    writer.raw('SET FOREIGN_KEY_CHECKS=1;')
    writer.raw('')

    # -------------------- Reference / Departments --------------------
    writer.insert_many('Insurance', ['Provider', 'Provider_Phone'], [(p, f'210{5000000+i:07d}') for i, p in enumerate(INSURANCES, 1)])

    bed_rows = []
    for dept in DEPARTMENTS:
        for n in range(1, 21):
            btype = 'ΜΕΘ' if dept == 'ΜΕΘ' or n <= 2 else rng.choice(['Μονόκλινο', 'Πολύκλινο'])
            status = rng.choices(['Διαθέσιμη', 'Κατειλημμένη', 'Υπό Συντήρηση'], weights=[8, 1, 1], k=1)[0]
            bed_rows.append((dept, n, btype, status))
    writer.insert_many('Bed', ['Department_Name', 'Bed_Number', 'Bed_Type', 'Status'], bed_rows)

    ken_rows = load_ken_reference(ken_path)
    diagnosis_rows = load_icd_reference(icd_paths)
    action_catalog = load_medical_action_catalog(action_paths)

    print(f'Reference KEN rows: {len(ken_rows)} from {ken_path}')
    print(f'Reference ICD rows: {len(diagnosis_rows)}')
    print(f'Reference medical action names: {len(action_catalog)}')
    if args.strict_reference:
        if not ken_path or len(ken_rows) <= len(KEN_ROWS):
            raise SystemExit('STRICT: Δεν φορτώθηκαν πλήρη ΚΕΝ από reference αρχείο.')
        if len(diagnosis_rows) <= len(DIAGNOSES):
            raise SystemExit('STRICT: Δεν φορτώθηκαν πλήρη ICD-10 από reference αρχείο.')
        if not action_catalog:
            raise SystemExit('STRICT: Δεν φορτώθηκαν πραγματικές ιατρικές πράξεις από xls/xlsx/csv.')

    writer.insert_many('KEN', ['KEN_Code', 'KEN_Cost', 'MDN'], ken_rows)
    writer.insert_many('Diagnosis', ['ICD_10_Code', 'Description'], diagnosis_rows)

    # Medicines / substances.
    # Reference data φαρμάκων:
    # - Με --ema-xlsx φορτώνουμε ΟΛΑ τα πραγματικά προϊόντα και τις δραστικές ουσίες του EMA Article 57.
    # - Χωρίς --ema-xlsx κρατάμε μόνο fallback demo rows, ώστε το script να τρέχει και μόνο του.
    medicines = load_ema_xlsx(ema_path, limit=None) if ema_path else FALLBACK_MEDICINES
    print(f'Reference EMA medicine rows: {len(medicines)} from {ema_path}')
    if ema_path is None:
        print('WARNING: Δεν δόθηκε EMA Article 57 αρχείο. Χρησιμοποιείται μικρό fallback sample· για πλήρη συμμόρφωση δώσε --ema-xlsx ή --strict-reference.')
    substance_id_by_name: dict[str, str] = {}
    active_rows = []
    med_rows = []
    comp_rows = []
    for idx, (ema, name, substances) in enumerate(medicines, 1):
        med_rows.append((ema, name[:200]))
        for sub in substances:
            sub_norm = sub.strip()[:200]
            if sub_norm not in substance_id_by_name:
                sid = f'SUB{len(substance_id_by_name) + 1:04d}'
                substance_id_by_name[sub_norm] = sid
                active_rows.append((sid, sub_norm))
            comp_rows.append((ema, substance_id_by_name[sub_norm]))
    writer.insert_many('Medicine', ['EMA_Code', 'Medicine_Name'], med_rows)
    writer.insert_many('Active_Substance', ['Substance_ID', 'Substance_Name'], active_rows)
    writer.insert_many('Medicine_Composition', ['EMA_Code', 'Substance_ID'], sorted(set(comp_rows)))

    # -------------------- Patients / Triage / Admissions --------------------
    patient_rows = []
    contact_rows = []
    patients: list[str] = []
    for i in range(1, 401):
        female = rng.random() < 0.52
        first, last = rand_name(rng, female=female)
        a = amka(400, i)
        patients.append(a)
        age = rng.randint(1, 92)
        age_month = 0 if age > 0 else rng.randint(0, 11)
        gender = 'Θήλυ' if female else 'Άρρεν'
        patient_rows.append((
            a, first, last, rng.choice(FATHER_NAMES), age, age_month, gender,
            round(rng.uniform(45, 115), 2), round(rng.uniform(1.45, 1.95), 2),
            f'Οδός Υγείας {i}', f'210{7000000+i:07d}', f'patient{i}@example.gr',
            rng.choice(['Ιδιωτικός Υπάλληλος', 'Δημόσιος Υπάλληλος', 'Φοιτητής', 'Συνταξιούχος', 'Άνεργος', 'Ελεύθερος Επαγγελματίας']),
            rng.choice(['Ελληνική', 'Κυπριακή', 'Αλβανική', 'Βουλγαρική', 'Ρουμανική']),
            rng.choice(INSURANCES)
        ))
        cf, cl = rand_name(rng)
        contact_rows.append((amka(500, i), a, f'69{30000000+i:08d}'[:10], cf, cl, f'contact{i}@example.gr'))
    writer.insert_many('Patient', ['Patient_AMKA', 'Patient_First_Name', 'Patient_Last_Name', 'Patient_Father_Name', 'Patient_Age', 'Age_Month', 'Patient_Gender', 'Patient_Weight', 'Patient_Height', 'Patient_Address', 'Patient_Phone_Number', 'Patient_Email', 'Patient_Profession', 'Patient_Nationality', 'Insurance_Provider'], patient_rows)
    writer.insert_many('Emergency_Contact', ['Contact_AMKA', 'Patient_AMKA', 'Contact_Phone_Number', 'Contact_First_Name', 'Contact_Last_Name', 'Contact_Email'], contact_rows)

    # Αλλεργίες: λίγες, και μετά στις συνταγογραφήσεις αποφεύγονται τα αντίστοιχα φάρμακα.
    substance_ids = list(substance_id_by_name.values())
    allergy_rows = []
    patient_allergies: dict[str, set[str]] = defaultdict(set)
    for p in rng.sample(patients, k=min(55, len(patients))):
        for sid in rng.sample(substance_ids, k=min(rng.choice([1, 1, 2]), len(substance_ids))):
            patient_allergies[p].add(sid)
            allergy_rows.append((p, sid))
    writer.insert_many('Patient_Allergy', ['Patient_AMKA', 'Substance_ID'], sorted(set(allergy_rows)))

    # Triage and Admission.
    triage_rows = []
    admission_rows = []
    admissions: list[dict[str, Any]] = []
    beds_by_dept = {dept: list(range(1, 21)) for dept in DEPARTMENTS}
    # Κρατάμε όλα τα διαστήματα ανά κλίνη ώστε να μη δημιουργούνται επικαλυπτόμενες νοσηλείες
    # στην ίδια συγκεκριμένη κλίνη, ακόμη και αν οι admissions παράγονται μη χρονολογικά.
    bed_intervals: dict[tuple[str, int], list[tuple[date, date]]] = defaultdict(list)

    def intervals_overlap(a0: date, a1: date, b0: date, b1: date) -> bool:
        # Θεωρούμε ότι η Release_Date είναι η ημέρα αποχώρησης, άρα το διάστημα είναι [Admission, Release).
        return a0 < b1 and b0 < a1

    def pick_non_overlapping_bed(dept: str, start: date, release: date | None) -> tuple[int, date, date | None]:
        end = release or (end_date + timedelta(days=365))
        candidate_beds = beds_by_dept[dept][:]
        rng.shuffle(candidate_beds)
        for bed_no in candidate_beds:
            intervals = bed_intervals[(dept, bed_no)]
            if all(not intervals_overlap(start, end, old_start, old_end) for old_start, old_end in intervals):
                intervals.append((start, end))
                return bed_no, start, release
        # Αν όλες οι κλίνες έχουν σύγκρουση, μετακινούμε ελάχιστα τη νοσηλεία μετά την
        # πρωιμότερη διαθέσιμη κλίνη. Αυτό είναι fallback ασφαλείας για πυκνά random δεδομένα.
        best_bed = min(candidate_beds, key=lambda b: max((iv[1] for iv in bed_intervals[(dept, b)]), default=start))
        next_start = max(iv[1] for iv in bed_intervals[(dept, best_bed)])
        duration = max(1, ((release or start) - start).days)
        if next_start > end_date - timedelta(days=duration):
            # Έσχατη λύση: κρατάμε την ημερομηνία και διαλέγουμε την κλίνη με το μικρότερο πλήθος
            # intervals. Στην κανονική κλίμακα των 300 κλινών δεν αναμένεται να ενεργοποιηθεί.
            best_bed = min(candidate_beds, key=lambda b: len(bed_intervals[(dept, b)]))
            bed_intervals[(dept, best_bed)].append((start, end))
            return best_bed, start, release
        new_start = next_start
        new_release = None if release is None else new_start + timedelta(days=duration)
        bed_intervals[(dept, best_bed)].append((new_start, new_release or (end_date + timedelta(days=365))))
        return best_bed, new_start, new_release

    ken_mdn = {code: mdn for code, _cost, mdn in ken_rows}
    ken_cost = {code: cost for code, cost, _mdn in ken_rows}

    # Seed για Q14: θέλουμε κατηγορίες ICD-10 που έχουν ίδιο πλήθος εισαγωγών
    # σε δύο συνεχόμενα έτη, με τουλάχιστον 5 περιστατικά ανά έτος.
    # Επειδή πολλά Q14 queries φιλτράρουν «τον τελευταίο χρόνο», κάνουμε το pattern
    # μέσα στο rolling last-12-month window: 5 admissions στο προηγούμενο έτος
    # μετά το last_year_start και 5 στο τρέχον έτος πριν/μέχρι end_date.
    def first_diag_with_prefix(prefix: str) -> str:
        for code, _desc in diagnosis_rows:
            if code.startswith(prefix):
                return code
        return diagnosis_rows[0][0]

    q14_seed: dict[int, tuple[date, str]] = {}
    q14_prefixes = ['I21', 'J18', 'K35']
    q14_diag_codes = [first_diag_with_prefix(p) for p in q14_prefixes]
    # Το δικό σου Q14 κάνει GROUP BY LEFT(ICD, 1), όχι LEFT(ICD, 3).
    # Άρα πρέπει να αποκλείσουμε ΟΛΕΣ τις random διαγνώσεις που αρχίζουν από I/J/K,
    # αλλιώς χαλάει η ισότητα των counts που φυτεύουμε.
    q14_excluded_categories = tuple(code[:1] for code in q14_diag_codes)
    diagnosis_rows_without_q14 = [d for d in diagnosis_rows if not d[0].startswith(q14_excluded_categories)] or diagnosis_rows

    last_year_start = end_date - timedelta(days=365)
    prev_year = end_date.year - 1
    curr_year = end_date.year
    prev_base = max(last_year_start + timedelta(days=3), date(prev_year, 6, 1))
    curr_base = date(curr_year, 1, 15)
    q14_id = 1
    for diag_code in q14_diag_codes:
        for n in range(5):
            q14_seed[q14_id] = (prev_base + timedelta(days=7 * n), diag_code)
            q14_id += 1
        for n in range(5):
            q14_seed[q14_id] = (min(end_date - timedelta(days=1), curr_base + timedelta(days=7 * n)), diag_code)
            q14_id += 1

    for adm_id in range(1, 1001):
        # Οι πρώτοι 5 ασθενείς έχουν >3 νοσηλείες στο ίδιο τμήμα για το Q3.
        if adm_id <= 20:
            p = patients[(adm_id - 1) // 4]
            dept = 'Παθολογία'
        # Οι επόμενες 400 νοσηλείες δίνουν τουλάχιστον 1 νοσηλεία σε κάθε ασθενή.
        elif adm_id <= 420:
            p = patients[(adm_id - 21) % len(patients)]
            dept = rng.choice(DEPARTMENTS)
        # Οι υπόλοιπες δημιουργούν πολλαπλές νοσηλείες, με bias σε μερικούς “frequent” ασθενείς.
        else:
            frequent_patients = patients[:80]
            p = rng.choice(frequent_patients) if rng.random() < 0.55 else rng.choice(patients)
            dept = rng.choice(DEPARTMENTS)

        ken = rng.choice(ken_rows)[0]
        mdn = ken_mdn[ken]
        base_cost = ken_cost[ken]

        # Κρατάμε περιθώριο ώστε αρκετές νοσηλείες να μπορούν όντως να ξεπεράσουν τη ΜΔΝ.
        adate = random_date(rng, start_date, max(start_date, end_date - timedelta(days=35)))
        if adm_id in q14_seed:
            q14_date, _q14_diag = q14_seed[adm_id]
            # Σταθερές ημερομηνίες μέσα στον τελευταίο χρόνο, ώστε τα δύο συνεχόμενα
            # έτη να έχουν ίδιο πλήθος ακόμη και αν το query κάνει filter last 12 months.
            adate = q14_date

        # Περίπου 8% ανοικτές νοσηλείες, μόνο κοντά στο τέλος.
        if rng.random() < 0.08 and adate > end_date - timedelta(days=20):
            rdate = None
            total_cost = None
        else:
            # Σκόπιμα βάζουμε περίπου 35% των ολοκληρωμένων νοσηλειών να ξεπερνούν τη ΜΔΝ,
            # ώστε τα queries για basic cost vs extra charge να επιστρέφουν μη μηδενικά αποτελέσματα.
            if rng.random() < 0.35:
                stay_days = mdn + rng.randint(1, 10)
            else:
                stay_days = rng.randint(1, max(1, mdn))
            rdate = min(adate + timedelta(days=stay_days), end_date)
            actual_days = max(1, (rdate - adate).days)
            extra_days = max(0, actual_days - mdn)
            # Ίδια λογική με την εκφώνηση: αναλογική πρόσθετη ημερήσια χρέωση μετά τη ΜΔΝ.
            # Αν το schema έχει trigger για Total_Cost, αυτή η τιμή είτε συμφωνεί είτε μπορεί να αγνοηθεί.
            total_cost = round(base_cost + extra_days * (base_cost / max(1, mdn)), 2)

        triage_dt = datetime.combine(adate, time(rng.randint(0, 23), rng.choice([0, 10, 20, 30, 40, 50])))
        nurse = rng.choice(nurse_members).amka
        wait = rng.randint(5, 240)
        urgency = rng.randint(1, 5)
        symptom = rng.choice(SYMPTOMS)
        # Για συμβατότητα με το trigger ins_triage: αρχικά μπαίνει Waiting_Minutes=NULL
        # και μετά γίνεται UPDATE με σειρά triage προτεραιότητας.
        triage_rows.append((adm_id, symptom, urgency, triage_dt, None, p, nurse, wait))
        if adm_id in q14_seed:
            diag_in = q14_seed[adm_id][1]
        else:
            # Αποφεύγουμε τα ίδια Q14 prefixes στο random κομμάτι, ώστε να μη χαλάσει
            # η ισότητα 5 vs 5 που φυτεύουμε για τον τελευταίο χρόνο.
            diag_in = rng.choice(diagnosis_rows_without_q14)[0]
        diag_out = rng.choice(diagnosis_rows)[0] if rdate else None
        bed, final_adate, final_rdate = pick_non_overlapping_bed(dept, adate, rdate)
        if final_adate != adate:
            # Αν χρειάστηκε μετακίνηση λόγω κλίνης, συγχρονίζουμε και το triage/admission.
            adate = final_adate
            rdate = final_rdate
            triage_dt = datetime.combine(adate, triage_dt.time())
            triage_rows[-1] = (adm_id, symptom, urgency, triage_dt, None, p, nurse, wait)
        admission_rows.append((adm_id, adate, rdate, total_cost, dept, bed, p, adm_id, ken, diag_in, diag_out))
        admissions.append({'id': adm_id, 'patient': p, 'dept': dept, 'adate': adate, 'rdate': rdate, 'ken': ken})

    # Επιπλέον περιστατικά triage που ΔΕΝ οδηγούν σε νοσηλεία, ώστε το Q15 να μη δίνει πάντα 100%.
    # Η εκφώνηση προβλέπει ότι κάποιοι ασθενείς παίρνουν οδηγίες και αποχωρούν.
    for extra_id in range(1001, 1151):
        p = rng.choice(patients)
        adate = random_date(rng, start_date, end_date)
        triage_dt = datetime.combine(adate, time(rng.randint(0, 23), rng.choice([0, 10, 20, 30, 40, 50])))
        nurse = rng.choice(nurse_members).amka
        wait = rng.randint(5, 180)
        urgency = rng.choices([1, 2, 3, 4, 5], weights=[1, 2, 4, 6, 7], k=1)[0]
        symptom = rng.choice(SYMPTOMS)
        triage_rows.append((extra_id, symptom, urgency, triage_dt, None, p, nurse, wait))

    triage_insert_rows = [r[:7] for r in triage_rows]
    triage_wait_updates = [(r[0], r[2], r[3], r[7]) for r in triage_rows]
    writer.insert_many('Triage', ['Triage_ID', 'Symptoms', 'Urgency_Level', 'Arrival_DateTime', 'Waiting_Minutes', 'Patient_AMKA', 'Nurse_AMKA'], triage_insert_rows)
    writer.raw('-- Ολοκλήρωση triage με σειρά προτεραιότητας, ώστε να περνάει το upd_triage trigger.')
    for tid, _urgency, _arrival, wait in sorted(triage_wait_updates, key=lambda x: (x[1], x[2], x[0])):
        writer.raw(f'UPDATE `Triage` SET `Waiting_Minutes` = {q(wait)} WHERE `Triage_ID` = {q(tid)};')
    writer.insert_many('Admission', ['AdmissionID', 'Admission_Date', 'Release_Date', 'Total_Cost', 'Department_Name', 'Bed_Number', 'Patient_AMKA', 'Triage_ID', 'KEN_Code', 'Admission_Diagnosis_ICD_10_Code', 'Release_Diagnosis_ICD_10_Code'], admission_rows)

    # -------------------- Shifts --------------------
    # Πλήρες 24/7 ημερολόγιο: κάθε ημέρα, κάθε τμήμα, και οι 3 βάρδιες.
    # Η στελέχωση γίνεται με κυκλική επιλογή προσωπικού και αυστηρό έλεγχο:
    # - Ιατροί έως 15 βάρδιες/μήνα
    # - Νοσηλευτές έως 20 βάρδιες/μήνα
    # - Διοικητικοί έως 25 βάρδιες/μήνα
    # - τουλάχιστον 8 ώρες ανάπαυσης
    # - τουλάχιστον 1 Επιμελητής Α ή Διευθυντής σε κάθε βάρδια
    staff_shift_rows = []
    month_counts: dict[tuple[str, int, int], int] = defaultdict(int)
    last_end_by_staff: dict[str, datetime] = {}
    night_streak_by_staff: dict[str, int] = defaultdict(int)
    last_night_date_by_staff: dict[str, date] = {}
    rr_pos: dict[tuple[str, int, int], int] = defaultdict(int)

    def can_take(member: StaffMember, s: datetime, e: datetime, limit: int, shift_type: str) -> bool:
        key = (member.amka, s.year, s.month)
        if month_counts[key] >= limit:
            return False
        prev_end = last_end_by_staff.get(member.amka)
        # Επειδή παράγουμε τις βάρδιες χρονολογικά, αρκεί να κοιτάμε την τελευταία λήξη.
        # Αν υπάρχει αργότερη/επικαλυπτόμενη βάρδια, απορρίπτεται από αυτό το check.
        if prev_end is not None and prev_end > s - timedelta(hours=8):
            return False
        if shift_type == 'Νυχτερινή':
            last_night = last_night_date_by_staff.get(member.amka)
            projected_streak = night_streak_by_staff[member.amka] + 1 if last_night == s.date() - timedelta(days=1) else 1
            if projected_streak > 3:
                return False
        return True

    def assign(group_name: str, cands: list[StaffMember], needed: int, s: datetime, e: datetime, limit: int, shift_type: str, chosen_so_far: set[str] | None = None) -> list[StaffMember]:
        chosen_so_far = chosen_so_far or set()
        key = (group_name, s.year, s.month)
        start_pos = rr_pos[key] % len(cands)
        chosen: list[StaffMember] = []
        checked = 0
        pos = start_pos
        while checked < len(cands) and len(chosen) < needed:
            m = cands[pos]
            if m.amka not in chosen_so_far and can_take(m, s, e, limit, shift_type):
                chosen.append(m)
                chosen_so_far.add(m.amka)
            pos = (pos + 1) % len(cands)
            checked += 1
        rr_pos[key] = pos
        if len(chosen) < needed:
            raise RuntimeError(
                f'Αδυναμία πλήρους στελέχωσης βάρδιας {s.date()} {group_name} {needed=} '
                f'με όριο {limit}. Αύξησε το πλήθος προσωπικού.'
            )
        return chosen

    senior_doctors = [d for d in doctor_members if d.rank in ('Επιμελητής Α', 'Διευθυντής')]
    additional_doctors = [d for d in doctor_members if d.rank == 'Επιμελητής Β']
    if len(senior_doctors) < 45:
        raise RuntimeError('Χρειάζονται τουλάχιστον 45 senior ιατροί για ημερήσια κάλυψη όλων των βαρδιών.')

    shift_rows = []
    for d in daterange(start_date, end_date):
        for dept in DEPARTMENTS:
            for stype, s_t, e_t in [
                ('Πρωινή', time(7, 0), time(15, 0)),
                ('Απογευματινή', time(15, 0), time(23, 0)),
                ('Νυχτερινή', time(23, 0), time(7, 0)),
            ]:
                sdt = datetime.combine(d, s_t)
                edt = datetime.combine(d if stype != 'Νυχτερινή' else d + timedelta(days=1), e_t)
                shift_rows.append((d, stype, sdt, edt, dept))

                picked: set[str] = set()
                senior = assign('senior_doctors', senior_doctors, 1, sdt, edt, 15, stype, picked)
                doctors = assign('additional_doctors', additional_doctors, 2, sdt, edt, 15, stype, picked)
                nurses = assign('nurses', nurse_members, 6, sdt, edt, 20, stype, picked)
                admins = assign('admins', management_members, 2, sdt, edt, 25, stype, picked)

                for m in senior + doctors + nurses + admins:
                    month_counts[(m.amka, sdt.year, sdt.month)] += 1
                    last_end_by_staff[m.amka] = edt
                    if stype == 'Νυχτερινή':
                        prev_night = last_night_date_by_staff.get(m.amka)
                        night_streak_by_staff[m.amka] = night_streak_by_staff[m.amka] + 1 if prev_night == sdt.date() - timedelta(days=1) else 1
                        last_night_date_by_staff[m.amka] = sdt.date()
                    staff_shift_rows.append((d, stype, dept, m.amka))

    writer.insert_many('Shift', ['Shift_Date', 'Shift_Type', 'Start_Time', 'End_Time', 'Department_Name'], shift_rows, batch=700)
    writer.insert_many('Shift_Staff', ['Shift_Date', 'Shift_Type', 'Department_Name', 'Staff_AMKA'], staff_shift_rows, batch=700)
    # Τα παραγόμενα δεδομένα είναι ήδη πλήρως στελεχωμένα. Αν θέλεις, μπορείς μετά τη φόρτωση
    # να τρέξεις CALL validate_shift_staffing(date, type, department) δειγματοληπτικά ή για όλες τις βάρδιες.

    # -------------------- Exams --------------------
    exam_rows = []
    for code in range(1, 241):
        adm = rng.choice(admissions)
        adate = adm['adate']
        max_date = adm['rdate'] or end_date
        edate = random_date(rng, adate, max_date)
        doc = rng.choice(doctor_members).amka
        etype = rng.choice(EXAM_TYPES)
        result = rng.choice(['Φυσιολογικό', 'Παθολογικό', 'Υπό παρακολούθηση', 'Θετικό', 'Αρνητικό'])
        unit = rng.choice(['mg/dL', 'mmol/L', 'IU/L', None])
        cost = round(rng.uniform(12, 350), 2)
        exam_rows.append((code, etype, edate, result, unit, cost, adm['id'], doc))
    writer.insert_many('Exam', ['Exam_Code', 'Exam_Type', 'Exam_Date', 'Exam_Result', 'Measurement_Unit', 'Exam_Cost', 'AdmissionID', 'Doctor_AMKA'], exam_rows)

    # -------------------- Operating Rooms / Medical Actions / Surgery --------------------
    room_rows = [(i, 'Χειρουργείο' if i <= 7 else 'Αίθουσα Επέμβασης') for i in range(1, 11)]
    writer.insert_many('Operating_Room', ['Room_Code', 'Room_Type'], room_rows)

    action_rows = []
    surgery_rows = []
    assistant_rows = []
    action_names = [
        'Αορτοστεφανιαία παράκαμψη', 'Λαπαροσκοπική χολοκυστεκτομή', 'Σκωληκοειδεκτομή',
        'Οστεοσύνθεση κατάγματος', 'Ενδοσκόπηση', 'Βιοψία', 'Καθετηριασμός', 'Παρακέντηση',
        'Θεραπευτική έγχυση', 'Αφαίρεση μορφώματος'
    ]
    num_medical_actions = max(180, len(action_catalog)) if action_catalog else 180
    surgeon_pool = [d for d in doctor_members if d.specialty in ('Χειρουργική', 'Ορθοπεδική', 'Καρδιολογία', 'Ουρολογία', 'Ογκολογία') or d.rank in ('Επιμελητής Α', 'Διευθυντής')]
    used_slots: set[tuple[int, datetime]] = set()
    completed_for_actions = [a for a in admissions if a['rdate'] is not None]
    total_days_for_actions = max(1, (end_date - start_date).days + 1)
    surgeon_next_start: dict[str, datetime] = {}
    assistant_next_start: dict[str, datetime] = {}
    # Εξασφαλίζουμε ότι το Q11 δεν βγαίνει κενό: δημιουργούμε αρκετές χειρουργικές
    # επεμβάσεις μέσα στο τρέχον έτος, με έναν top surgeon και άλλους με τουλάχιστον
    # 5 λιγότερες επεμβάσεις από αυτόν.
    current_year = end_date.year
    current_year_start = date(current_year, 1, 1)
    seeded_surgery_counts: dict[str, int] = {}
    top_surgeon = surgeon_pool[0].amka
    other_surgeons_for_q11 = [d.amka for d in surgeon_pool[1:16]]
    q11_plan = [(top_surgeon, 14)] + [(s, rng.randint(1, 7)) for s in other_surgeons_for_q11]
    q11_surgery_owner: dict[int, str] = {}
    q11_code = 1
    for surgeon_amka, cnt in q11_plan:
        for _ in range(cnt):
            q11_surgery_owner[q11_code] = surgeon_amka
            seeded_surgery_counts[surgeon_amka] = seeded_surgery_counts.get(surgeon_amka, 0) + 1
            q11_code += 1

    for code in range(1, num_medical_actions + 1):
        adm = completed_for_actions[(code - 1) % len(completed_for_actions)]
        catalog_entry = action_catalog[(code - 1) % len(action_catalog)] if action_catalog else None
        is_surgery = code in q11_surgery_owner or code <= 150 or (catalog_entry is not None and catalog_entry['type'] == 'Χειρουργική')
        atype = 'Χειρουργική' if is_surgery else (catalog_entry['type'] if catalog_entry else rng.choice(['Διαγνωστική', 'Θεραπευτική']))
        action_name = catalog_entry['name'] if catalog_entry else rng.choice(action_names)
        # Κρατάμε duration <= 50' ώστε τα hourly room slots να μην επικαλύπτονται.
        duration = rng.randint(35, 50) if is_surgery else rng.randint(15, 45)
        cost = round(rng.uniform(300, 6000) if is_surgery else rng.uniform(80, 900), 2)
        main = q11_surgery_owner.get(code, rng.choice(surgeon_pool).amka) if is_surgery else None

        # Deterministic hourly slots: κάθε room έχει το πολύ μία πράξη ανά ώρα.
        # Για χειρουργικές πράξεις ελέγχουμε επιπλέον και διαθεσιμότητα main surgeon.
        slot = code - 1
        while True:
            day_offset = (slot // 100) % total_days_for_actions
            within_day = slot % 100
            room = (within_day % 10) + 1
            hour = 7 + (within_day // 10)  # 07:00 έως 16:00
            start_dt = datetime.combine(start_date + timedelta(days=day_offset), time(hour, 0))
            if code in q11_surgery_owner:
                # Για το Q11 κρατάμε τις seeded επεμβάσεις στο τρέχον έτος, αλλά το slot
                # βασίζεται στο μεταβαλλόμενο `slot`, ώστε αν υπάρχει σύγκρουση να μπορεί
                # να μετακινηθεί και να μη δημιουργηθεί infinite loop.
                q11_slot = max(0, slot - 1)
                q11_day_offset = q11_slot // 10
                q11_hour = 8 + (q11_slot % 10)
                q11_date = current_year_start + timedelta(days=q11_day_offset)
                if q11_date > end_date:
                    q11_date = end_date
                start_dt = datetime.combine(q11_date, time(q11_hour, 0))
                room = (q11_slot % 10) + 1
            end_dt = start_dt + timedelta(minutes=duration)
            room_ok = (room, start_dt) not in used_slots
            # Για μη-seeded χειρουργικές επιλέγουμε τον κύριο χειρουργό με βάση το τρέχον slot.
            # Έτσι δεν κολλάμε σε γιατρό που είναι απασχολημένος μέχρι το τέλος του διαθέσιμου εύρους.
            if is_surgery and code not in q11_surgery_owner:
                available_main_surgeons = [m.amka for m in surgeon_pool if start_dt >= surgeon_next_start.get(m.amka, datetime.min)]
                if available_main_surgeons:
                    main = rng.choice(available_main_surgeons)
                else:
                    slot += 1
                    continue
            surgeon_ok = True
            assistants_for_action: list[str] = []
            if main is not None:
                surgeon_ok = start_dt >= surgeon_next_start.get(main, datetime.min)
                # Βοηθοί διαθέσιμοι στο ίδιο χρονικό slot. Χρησιμοποιούμε νοσηλευτές ως βοηθούς,
                # ώστε να αποφεύγεται και η έμμεση σύγκρουση με κύριους χειρουργούς.
                available_assistants = [
                    m.amka for m in nurse_members
                    if m.amka != main and start_dt >= assistant_next_start.get(m.amka, datetime.min)
                    and start_dt >= surgeon_next_start.get(m.amka, datetime.min)
                ]
                if len(available_assistants) >= 2:
                    assistants_for_action = rng.sample(available_assistants, k=2)
                else:
                    surgeon_ok = False
            if room_ok and surgeon_ok:
                break
            slot += 1
        used_slots.add((room, start_dt))
        if main is not None:
            surgeon_next_start[main] = end_dt
        action_rows.append((code, action_name[:400], atype, start_dt, duration, cost, adm['id'], room))
        if is_surgery:
            surgery_rows.append((rng.choice(['Τακτική', 'Επείγουσα', 'Ελάχιστα Επεμβατική']), code, main))
            for aa in assistants_for_action:
                assistant_next_start[aa] = end_dt
                assistant_rows.append((code, aa))
    writer.insert_many('Medical_Action', ['Action_Code', 'Action_Name', 'Action_Type', 'Action_Start', 'Action_Duration', 'Action_Cost', 'AdmissionID', 'Operating_Room_Code'], action_rows)
    writer.insert_many('Surgery', ['Surgery_Type', 'Action_Code', 'Main_Surgeon_AMKA'], surgery_rows)
    writer.insert_many('Surgery_Assistant', ['Surgery_Action_Code', 'Assistant_AMKA'], assistant_rows)

    # -------------------- Prescriptions --------------------
    med_subs: dict[str, set[str]] = defaultdict(set)
    for ema, sid in comp_rows:
        med_subs[ema].add(sid)
    medicine_codes = [m[0] for m in med_rows]
    popular_docs = [d.amka for d in rng.sample(doctor_members, 6)]

    prescription_rows = []
    prescribed_by_admission: dict[int, set[str]] = defaultdict(set)
    # Κρατάμε και τα δύο uniqueness levels:
    # 1) αυτό που επιβάλλει το schema: (Doctor, Admission, Medicine, Start_Date)
    # 2) αυτό που ζητάει ρητά η εκφώνηση: (Doctor, Patient, Medicine, Start_Date)
    prescription_schema_keys: set[tuple[str, int, str, date]] = set()
    prescription_patient_keys: set[tuple[str, str, str, date]] = set()
    pres_id = 1
    completed_adms = [a for a in admissions if a['rdate'] is not None]
    attempts = 0
    while pres_id <= 650 and attempts < 5000:
        attempts += 1
        adm = rng.choice(completed_adms)
        patient = adm['patient']
        # Κάποιοι γιατροί έχουν πληθώρα αξιολογήσεων/συνταγογραφήσεων.
        doc = rng.choice(popular_docs) if rng.random() < 0.38 else rng.choice(doctor_members).amka
        blocked_substances = patient_allergies.get(patient, set())
        # Με πλήρες EMA υπάρχουν 160k+ φάρμακα. Μην χτίζεις safe_meds list σε κάθε
        # prescription, γιατί γίνεται άσκοπα αργό. Κάνουμε random rejection sampling.
        med = None
        for _ in range(200):
            candidate_med = rng.choice(medicine_codes)
            if not (med_subs[candidate_med] & blocked_substances):
                med = candidate_med
                break
        if med is None:
            safe_meds = [m for m in medicine_codes if not (med_subs[m] & blocked_substances)]
            if not safe_meds:
                continue
            med = rng.choice(safe_meds)
        sd = random_date(rng, adm['adate'], adm['rdate'])
        ed = min(sd + timedelta(days=rng.randint(1, 14)), adm['rdate'])
        schema_key = (doc, adm['id'], med, sd)
        patient_key = (doc, patient, med, sd)
        if schema_key in prescription_schema_keys or patient_key in prescription_patient_keys:
            continue
        prescription_schema_keys.add(schema_key)
        prescription_patient_keys.add(patient_key)
        prescription_rows.append((pres_id, sd, ed, rng.choice([1, 1, 2, 500, 850]), rng.choice([1, 2, 3]), doc, med, adm['id']))
        prescribed_by_admission[adm['id']].add(doc)
        pres_id += 1
    writer.insert_many('Prescription', ['Prescription_ID', 'Start_Date', 'End_Date', 'Dosage', 'Frequency', 'Doctor_AMKA', 'EMA_Code', 'AdmissionID'], prescription_rows)

    # -------------------- Evaluations --------------------
    eval_rows = []
    doctor_eval_rows = []
    for adm in completed_adms:
        if rng.random() > 0.82:
            continue
        # Ρεαλιστική διασπορά Likert: κυρίως 3-5, αλλά όχι όλα τέλεια.
        nursing = rng.choices([1, 2, 3, 4, 5], weights=[1, 4, 18, 38, 29], k=1)[0]
        clean = rng.choices([1, 2, 3, 4, 5], weights=[2, 6, 22, 35, 25], k=1)[0]
        food = rng.choices([1, 2, 3, 4, 5], weights=[5, 12, 30, 30, 13], k=1)[0]
        overall = max(1, min(5, round((nursing + clean + food) / 3 + rng.choice([-1, 0, 0, 1]))))
        eval_rows.append((nursing, clean, food, overall, adm['id']))
        for doc in prescribed_by_admission.get(adm['id'], set()):
            dq = rng.choices([1, 2, 3, 4, 5], weights=[2, 5, 18, 40, 35], k=1)[0]
            doctor_eval_rows.append((adm['id'], doc, dq))
    writer.insert_many('Evaluation', ['Nursing_Quality', 'Cleanliness', 'Food', 'Overall_Experience', 'AdmissionID'], eval_rows)
    writer.insert_many('Doctor_Evaluation', ['AdmissionID', 'Doctor_AMKA', 'Doctor_Quality'], doctor_eval_rows)

    # -------------------- Entity images --------------------
    # Κάλυψη όλων των entity types που επιτρέπει το schema.
    image_rows = []
    img_id = 1

    def add_image(entity_type: str, entity_key: Any, description: str) -> None:
        nonlocal img_id
        safe_key = str(entity_key).replace(' ', '_').replace('/', '_')[:80]
        image_rows.append((img_id, entity_type, str(entity_key)[:100], f'https://example.org/images/{entity_type.lower()}/{safe_key}.jpg', description[:255]))
        img_id += 1

    for dept in DEPARTMENTS:
        add_image('Department', dept, f'Εικόνα τμήματος {dept}')
    for d in doctor_members:
        add_image('Doctor', d.amka, f'Φωτογραφία ιατρού {d.first} {d.last}')
    for n in nurse_members:
        add_image('Nurse', n.amka, f'Φωτογραφία νοσηλευτικού προσωπικού {n.first} {n.last}')
    for m in management_members:
        add_image('Management', m.amka, f'Φωτογραφία διοικητικού προσωπικού {m.first} {m.last}')
    for p in patients:
        add_image('Patient', p, f'Ενδεικτική εικόνα/φάκελος ασθενούς {p}')
    for dept, bed_no, btype, _status in bed_rows:
        add_image('Bed', f'{dept}#{bed_no}', f'Εικόνα κλίνης {bed_no} στο τμήμα {dept} ({btype})')
    for ema, name in med_rows:
        add_image('Medicine', ema, f'Εικόνα φαρμάκου {name}')
    for room, rtype in room_rows:
        add_image('Operating_Room', room, f'Εικόνα {rtype} {room}')
    for code, name, *_rest in action_rows:
        add_image('Medical_Action', code, f'Εικόνα/εικονίδιο ιατρικής πράξης: {name}')
    for code, etype, *_rest in exam_rows:
        add_image('Exam', code, f'Εικόνα/εικονίδιο εξέτασης {etype}')
    for _stype, action_code, _main in surgery_rows:
        add_image('Surgery', action_code, f'Εικόνα/εικονίδιο χειρουργικής επέμβασης {action_code}')
    for code, desc in diagnosis_rows:
        add_image('Diagnosis', code, f'Εικόνα/εικονίδιο διάγνωσης ICD-10: {desc}')
    for code, cost, mdn in ken_rows:
        add_image('KEN', code, f'Εικόνα/εικονίδιο ΚΕΝ {code}, κόστος {cost}, ΜΔΝ {mdn}')

    writer.insert_many('Entity_Image', ['Image_ID', 'Entity_Type', 'Entity_Key', 'Image_URL', 'Description'], image_rows, batch=700)

    writer.raw('')
    writer.raw('SET UNIQUE_CHECKS=@OLD_UNIQUE_CHECKS;')
    writer.raw('SET FOREIGN_KEY_CHECKS=@OLD_FOREIGN_KEY_CHECKS;')
    writer.raw('')
    writer.raw('-- Summary:')
    writer.raw(f'-- Doctors: {len(doctor_rows)}')
    writer.raw(f'-- Nurses: {len(nurse_rows)}')
    writer.raw(f'-- Management: {len(management_rows)}')
    writer.raw(f'-- Patients: {len(patient_rows)}')
    writer.raw(f'-- Triage cases: {len(triage_rows)}')
    writer.raw(f'-- Admissions: {len(admission_rows)}')
    writer.raw(f'-- Triage without admission: {len(triage_rows) - len(admission_rows)}')
    writer.raw(f'-- Shifts: {len(shift_rows)}')
    writer.raw(f'-- Covered Shift_Staff rows: {len(staff_shift_rows)}')
    writer.raw(f'-- Prescriptions: {len(prescription_rows)}')
    writer.raw(f'-- Exams: {len(exam_rows)}')
    writer.raw(f'-- Medical actions: {len(action_rows)}')
    writer.raw(f'-- Surgeries: {len(surgery_rows)}')
    writer.raw(f'-- Evaluations: {len(eval_rows)}')
    writer.raw(f'-- Doctor evaluations: {len(doctor_eval_rows)}')

    # -------------------- Generator self-checks --------------------
    # Αυτά δεν είναι SQL comments· είναι πραγματικοί έλεγχοι του generator.
    # Αν κάποια μελλοντική αλλαγή ξαναχαλάσει το triage/admission ή τα δεδομένα,
    # το script σταματάει αντί να γράψει προβληματικό load.sql.
    triage_wait_by_id = {r[0]: r[7] for r in triage_rows}
    if len(triage_wait_by_id) != len(triage_rows):
        raise RuntimeError('Self-check failed: duplicate Triage_ID generated.')
    if any(wait is None for wait in triage_wait_by_id.values()):
        raise RuntimeError('Self-check failed: generated triage without final Waiting_Minutes.')
    if any(r[7] not in triage_wait_by_id for r in admission_rows):
        raise RuntimeError('Self-check failed: Admission references missing Triage_ID.')

    # Προσομοίωση του upd_triage trigger: τα UPDATE πρέπει να ολοκληρώνουν πρώτα
    # υψηλότερη προτεραιότητα και, στην ίδια προτεραιότητα, FIFO βάσει Arrival_DateTime.
    triage_priority = {r[0]: (r[2], r[3]) for r in triage_rows}
    completed_triages: set[int] = set()
    for tid, _urgency, _arrival, _wait in sorted(triage_wait_updates, key=lambda x: (x[1], x[2], x[0])):
        urgency, arrival = triage_priority[tid]
        has_pending_priority = any(
            other_tid != tid
            and other_tid not in completed_triages
            and (other_urgency < urgency or (other_urgency == urgency and other_arrival < arrival))
            for other_tid, (other_urgency, other_arrival) in triage_priority.items()
        )
        if has_pending_priority:
            raise RuntimeError('Self-check failed: triage UPDATE order violates priority/FIFO trigger.')
        completed_triages.add(tid)

    if len(completed_triages) != len(triage_rows):
        raise RuntimeError('Self-check failed: not all triage rows receive Waiting_Minutes UPDATE.')

    admission_by_id = {a['id']: a for a in admissions}
    patient_natural_keys = [
        (r[5], admission_by_id[r[7]]['patient'], r[6], r[1])
        for r in prescription_rows
    ]
    if len(patient_natural_keys) != len(set(patient_natural_keys)):
        raise RuntimeError('Self-check failed: duplicate prescription natural key (Doctor, Patient, Medicine, Start_Date).')

    allergy_conflicts = []
    for r in prescription_rows:
        patient = admission_by_id[r[7]]['patient']
        if med_subs[r[6]] & patient_allergies.get(patient, set()):
            allergy_conflicts.append(r[0])
    if allergy_conflicts:
        raise RuntimeError(f'Self-check failed: prescription allergy conflicts: {allergy_conflicts[:5]}')

    completed_admission_ids = {a['id'] for a in admissions if a['rdate'] is not None}
    if any(r[-1] not in completed_admission_ids for r in eval_rows):
        raise RuntimeError('Self-check failed: evaluation for non-completed admission.')
    prescribed_pairs = {(r[7], r[5]) for r in prescription_rows}
    if any((adm_id, doctor) not in prescribed_pairs for adm_id, doctor, _score in doctor_eval_rows):
        raise RuntimeError('Self-check failed: doctor evaluation without same-admission prescription.')

    Path(args.output).write_text(writer.text(), encoding='utf-8')
    print(f'Wrote {args.output}')


if __name__ == '__main__':
    main()
